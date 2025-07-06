from flask import Flask, render_template, redirect, url_for, request
import sqlite3
import pandas as pd # 引用套件並縮寫為 pd
import json
import os
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from bs4 import BeautifulSoup
import pymysql

# 資料庫參數設定
db_settings = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "gimy0710",
    "db": "my_db",
    "charset": "utf8"
}

global_value = 0
def to_do():
    global_value += 1

def df_transform(df):
    data = []
    for row in df.values:
        a = list(row[0:2])
        for i in range(2,len(row)):
            #a = [row[0],row[1],df.columns[i],row[i]]
            b = a + [df.columns[i]]
            b.append(row[i])
            data.append(b)
    df = pd.DataFrame(data,columns=['STAGE','ITEM','PERIOD','VAL'])
    return df
def get_df_from_mysql(SQL):
    cn = pymysql.connect(**db_settings)
    df = pd.read_sql(SQL, cn)
    cn.close()
    return df
def get_df_from_sqlite(SQL):
    cn = sqlite3.connect('my_db.db')
    df = pd.read_sql(SQL, cn)
    cn.close()
    return df

app = Flask(__name__)

@app.route("/")
def home():
    return "HELLO222"
@app.route("/home2")
def home2():
    cn = sqlite3.connect('my_db.db')
    SQL = "select *,(case when grade=0 then 'pink' else 'white' end) as COLOR from books order by pass_datetime desc"
    df = pd.read_sql(SQL, cn)
    SQL = "select strftime('%Y%W',pass_datetime),sum(grade) from books group by strftime('%Y%W',pass_datetime)"
    df2 = pd.read_sql(SQL, cn)
    cn.close()
    jf = df.to_json(orient="records")
    html = '<a href="/static/index.html">index</a>'
    html += ' | <a href="/create_table">create table</a>'
    html += ' | <a href="/insert">資料插入</a>'
    html += ' | <a href="/web_scraper">爬蟲</a>'
    html += '<h3>' + jf
    html += '</h3>'
    #return html
    jf2 = df2.to_json(orient="records")
    return render_template('index.html',jf=jf,jf2=jf2)

@app.route("/mysql_query", methods=["POST", "GET"])
def mysql_query():
    SQL = "select * from cost_course where MFG_DAY='2023/11/04' and STAGE='A'"
    if request.method == "POST":
        SQL = request.form["sql"]
    df = get_df_from_mysql(SQL)
    jf = df.to_json(orient="records", date_format="iso")
    return jf
    
@app.route("/json_query")
def json_query():
    cn = sqlite3.connect('my_db.db')
    SQL = "select * from expense"
    SQL = "select * from books"
    df = pd.read_sql(SQL, cn)
    cn.close()
    jf = df.to_json(orient="records")
    return jf
    
    
@app.route("/create_table")
def create_table():
    cn = sqlite3.connect('my_db.db')
    SQL = "create table IF NOT EXISTS expense (PERIOD date, CLASS2 char(10), EXPENSE int)"
    SQL = "create table IF NOT EXISTS books (id char(50) primary key,book_name char(50),result char(10),grade int, pass_datetime datetime)"
    cn.execute(SQL)
    cn.close()
    return '<a href="/">expense 表格建立完成</a>'

@app.route("/insert")
def insert():
    cn = sqlite3.connect('my_db.db')
    SQL = "insert into expense (PERIOD,CLASS2,EXPENSE) values ('2023-08-01','維護費用',10000)"
    cn.execute(SQL)
    cn.commit()
    cn.close()
    return '<a href="/">資料插入一筆</a>'

@app.route("/add_book")
def add_book():
    cn = sqlite3.connect('my_db.db')
    SQL = "insert OR IGNORE into books (id) values ('test')"
    cn.execute(SQL)
    cn.commit()
    cn.close()
    return '<a href="/static/index.html">back</a>'

@app.route("/del_book")
def del_book():
    cn = sqlite3.connect('my_db.db')
    SQL = "delete from books where  id='test'"
    cn.execute(SQL)
    cn.commit()
    cn.close()
    return '<a href="/static/index.html">back</a>'

@app.route('/run_def/<s>')
def run_def(s):
    return '<a href="/">back</a>'

@app.route('/excel/<sheet_name>')
def excel(sheet_name):
    path = r'D:\doc\test.xlsx'
    wb = openpyxl.load_workbook(path)
    #sheet_name = wb.sheetnames[0]
    #s = wb[sheet_name]
    s = wb[wb.sheetnames[0]]
    s['A10'].value = sheet_name
    wb.save(path)
    wb.close()
    os.system(path)
    #return f'<a href="/excel/bbb">Hello</a> {sheet_name}'
    return '<a href="/excel/aaa">aaa</a> <a href="/excel/bbb">bbb</a> '

@app.route("/web_scraper")
def web_scraper():
    driver = webdriver.Edge()
    url = "https://read.tc.edu.tw/member_login_new.php"
    driver.get(url) # 更改網址以前往不同網頁
    a = driver.find_element('id', 'student')
    a.click()
    a = driver.find_element(By.XPATH, "//input[@type='radio' and @value='s_uname']")
    a.click()
    a = driver.find_element(By.NAME, "stu_username")
    a.send_keys('ltestyc111030604')
    a = driver.find_element(By.NAME, "password")
    a.send_keys('71565')
    a = driver.find_element(By.NAME, "submit_item")
    a.click()
    url = "https://read.tc.edu.tw/stu_account_number_1.php"
    driver.get(url) # 更改網址以前往不同網頁
    a = driver.find_element(By.NAME, "perpage")
    Select(a).select_by_index(2)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    driver.close() # 關閉瀏覽器視窗
    
    cn = sqlite3.connect('my_db.db')
    tr = soup.find_all('tr')
    for i in tr:
        td = i.find_all('td')
        if len(td) > 5:
            print(td[1].text,td[2].text,td[3].text,td[4].text,td[5].text)
            result = td[5].text
            grade = td[6].text
            SQL = "INSERT OR IGNORE INTO books (id) VALUES ('" + td[1].text + "')"
            cn.execute(SQL)
            pass_datetime = td[7].text.split('~')[0]
            SQL = "update books set book_name='" + td[2].text + "',result='" + result + "',grade='" + grade + "',pass_datetime='" + pass_datetime + "' where id='" + td[1].text + "'"
            cn.execute(SQL)
    cn.commit()
    cn.close()
    return '<a href="/">爬蟲執行完成</a>'
    
@app.route("/hello")
def hello():
    return "Hello World! This is Hello Page "

@app.route('/memo/', methods=['GET','POST'])
def memo():
    cn = sqlite3.connect("my_db.db")
    if 'MSN' in request.form:
        memo = request.form['MEMO'].replace("'","''")
        SQL = "update memo set "
        SQL += "MEMO='" + memo + "'"
        SQL += " where SN=" + request.form['MSN']
        if 'ACT' in request.form:
            SQL = "delete from memo where SN=" + request.form['MSN']
        cn.execute(SQL)
        cn.commit()
    elif 'MEMO' in request.form:
        if request.form['MEMO']:
            memo = request.form['MEMO'].replace("'","''")
            SQL = "insert into memo (MEMO) values ('" + memo + "')"
            cn.execute(SQL)
            cn.commit()
    SQL = "select * from memo order by MEMO"
    df = pd.read_sql(SQL, cn)
    jf = df.to_json(orient="records")
    cn.close()
    return render_template('memo.html',myvar=jf)

@app.route('/meeting/', methods=['GET','POST'])
def meeting():
    cn = sqlite3.connect("my_db.db")
    if 'MSN' in request.form:
        SQL = "update meeting_minutes set "
        SQL += "ITEM='" + request.form['ITEM'] + "'"
        SQL += ",OWNER='" + request.form['OWNER'] + "'"
        SQL += ",DUE_DATE='" + request.form['DUE_DATE'] + "' where SN=" + request.form['MSN']
        cn.execute(SQL)
        cn.commit()
    elif 'ITEM' in request.form:
        if request.form['ITEM']:
            SQL = "insert into meeting_minutes (ITEM) values ('" + request.form['ITEM'] + "')"
            cn.execute(SQL)
            cn.commit()
    SQL = "select * from meeting_minutes where BELONG is null"
    df = pd.read_sql(SQL, cn)
    jf = df.to_json(orient="records")
    cn.close()
    return render_template('meeting.html',myvar=jf)

@app.route('/meeting_minutes/<sn>', methods=['GET','POST'])
def meeting_minutes(sn):
    #sn = request.args.get('sn')
    cn = sqlite3.connect("my_db.db")
    SQL = "select * from meeting_minutes where SN=" + sn
    df = pd.read_sql(SQL, cn)
    title = '會議名稱:' + df["ITEM"][0] + ' 主持人:' + df["OWNER"][0] + ' 日期:' + df["DUE_DATE"][0]
    '''
    if request.method == 'POST':
        if request.form['name']:
            SQL = "insert into meeting_minutes (ITEM,OWNER,DUE_DATE,BELONG) values ('" + request.form['name'] + "'"
            SQL += ",'" + request.form['owner'] + "'"
            SQL += ",'" + request.form['date'] + "'," + sn + ")"
            cn.execute(SQL)
            cn.commit()
    '''
    if 'MSN' in request.form:
        SQL = "update meeting_minutes set ITEM='" + request.form['ITEM'] + "'"
        SQL += ",OWNER='" + request.form['OWNER'] + "'"
        SQL += ",DUE_DATE='" + request.form['DUE_DATE'] + "'"
        SQL += " where SN=" + request.form['MSN']
        cn.execute(SQL)
        cn.commit()
    elif 'ITEM' in request.form:
        SQL = "insert into meeting_minutes (ITEM,OWNER,DUE_DATE,BELONG) values ('" + request.form['ITEM'] + "'"
        SQL += ",'" + request.form['OWNER'] + "'"
        SQL += ",'" + request.form['DUE_DATE'] + "'," + sn + ")"
        cn.execute(SQL)
        cn.commit()        
    SQL = "select * from meeting_minutes where BELONG=" + sn
    df = pd.read_sql(SQL, cn)
    jf = df.to_json(orient="records")
    cn.close()
    return render_template('meeting_minutes.html',title=title,myvar=jf,sn=sn)
@app.route('/init/')
def init():
    cn = sqlite3.connect("my_db.db")
    #SQL = "create table collab (SN INTEGER PRIMARY KEY AUTOINCREMENT, LM_TIME datetime default current_timestamp)"
    #SQL = "insert or replace into collab (SN) values (2)"
    #SQL = "alter table meeting_minutes add column OWNER text after ITEM"
    #SQL = "create table IF NOT EXISTS meeting_minutes (SN INTEGER PRIMARY KEY AUTOINCREMENT,ITEM TEXT,OWNER text,DUE_DATE date,BELONG INTEGER, LM_TIME datetime default current_timestamp)"
    SQL = "create table IF NOT EXISTS memo (SN INTEGER PRIMARY KEY AUTOINCREMENT,MEMO TEXT,LM_TIME datetime default current_timestamp)"
    SQL = "create table books (id text PRIMARY KEY,book_name text,result text,grade INTERGER,pass_datetime datetime,LM_TIME datetime default current_timestamp)"
    #cn.execute("drop table menu")
    SQL = "create table IF NOT EXISTS menu (SN INTEGER PRIMARY KEY AUTOINCREMENT,LINK_NAME TEXT,LINK_ADDR text,BELONG INTEGER,LM_TIME datetime default current_timestamp)"
    cn.execute(SQL)
    cn.commit()
    cn.close()
    return '<a href="/meeting/">完成資料表建立</a>'
    
@app.route('/expense/')
def expense():
    cn = sqlite3.connect("my_db.db")
    SQL = "select * from expense where MFG_DAY < '2023-01-03'"
    SQL = "select MFG_DAY,sum(EXPENSE) as EXPENSE from expense group by MFG_DAY"
    df = pd.read_sql(SQL, cn)
    jf = df.to_json(orient="records")

    SQL = "select DEPT"
    for i in range(1,20):
        SQL += ",sum(case when MFG_DAY='2023-01-" + "{0:02d}".format(i) + "' then EXPENSE else 0 end) as '2023/01/" + "{0:02d}".format(i) + "'"
    SQL += " from expense where MFG_DAY <= '2023-01-02' group by DEPT"
    df = pd.read_sql(SQL, cn)
    jf2 = df.to_json(orient="records")
    
    cn.close()
    return render_template('expense.html',jf=jf,jf2=jf2)


@app.route('/menu/', methods=['GET','POST'])
def menu():
    cn = sqlite3.connect("my_db.db")
    if 'MSN' in request.form:
        SQL = "update menu set "
        SQL += "LINK_NAME='" + request.form['LINK_NAME'] + "'"
        SQL += ",LINK_ADDR='" + request.form['LINK_ADDR'] + "'"
        SQL += ",BELONG=0 where SN=" + request.form['MSN']
        cn.execute(SQL)
        cn.commit()
    elif 'BELONG' in request.form:
        if request.form['BELONG']:
            SQL = "insert into menu (LINK_NAME,LINK_ADDR,BELONG) values ('" + request.form['LINK_NAME'] + "','" + request.form['LINK_ADDR'] + "'," + request.form['BELONG'] + ")"
            cn.execute(SQL)
            cn.commit()
    elif 'LINK_NAME' in request.form:
        if request.form['LINK_NAME']:
            SQL = "insert into menu (LINK_NAME,LINK_ADDR) values ('" + request.form['LINK_NAME'] + "','" + request.form['LINK_ADDR'] + "')"
            cn.execute(SQL)
            cn.commit()
    SQL = "select * from menu"
    df = pd.read_sql(SQL, cn)
    jf = df.to_json(orient="records")
    cn.close()
    return render_template('menu.html',json=jf)

@app.route('/login_test/', methods=['GET','POST'])
def login_test():
    #name =flask.Request.remote_user.name
    remote_user = str(request.remote_user)
    ip = request.remote_addr + remote_user
    return ip

@app.route("/kanban/")
def kanban():
    SQL = "select strftime('%Y-%m-%d',pass_datetime) as 日期,sum(grade) as grade,count(*) as 次數,sum(case when grade > 0 then 1 else 0 end) as 成功,sum(case when grade > 0 then 1 else 0 end)*100/sum(1) as 成功率 from books where pass_datetime >= '2024-03-01' group by strftime('%Y-%m-%d',pass_datetime)"
    df = get_df_from_sqlite(SQL)
    jf = df.to_json(orient="records")
    return render_template('kanban.html',jf=jf)

    
if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)
    
    
'''
from flask import Flask
import webbrowser

app = Flask(__name__)

@app.route("/")
def hello():
    a = "Hello Flask!"
    return a
    

if __name__ == "__main__":
    app.run()
    url='http://127.0.0.1:5000/'
    webbrowser.get('windows-default').open_new(url)
'''
